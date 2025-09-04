#!/usr/bin/env python3

import csv
import re
import tempfile
from typing import Callable, Dict, List, Optional, Type

import sys
from collections import defaultdict
import argparse
import logging

logging.basicConfig(level=logging.INFO)

logger = logging.getLogger(__name__)


def convert_xlsx_to_csv(file_path: str, start_marker="U Location") -> str:
    try:
        import openpyxl
    except ImportError:
        raise ValueError("openpyxl not available - please run 'pip install openpyxl' to parse excel files")
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    start_row = 0
    if start_marker:
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):  # start=1 to get 1-based index
            if row[0] == start_marker:
                start_row = idx
                break
        else:
            raise ValueError(f"Marker '{start_marker}' not found in {file_path}.")

    file_name = file_path.rsplit("/", 1)[-1]
    file_name = file_name.rsplit(".", 1)[0]

    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", prefix=file_name, delete=False) as temp_file:
        csv_writer = csv.writer(temp_file)
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            csv_writer.writerow(row)
        return temp_file.name


def format_mac(mac: str) -> str:
    clean = re.sub(r'[^0-9A-Fa-f]', '', mac).lower()
    if len(clean) != 12:
        return ""
    return ':'.join(clean[i:i + 2] for i in range(0, 12, 2))


def fmt_dict(d: dict) -> str:
    return "\n".join([f"   {k}: {v}" for k, v in d.items() if v]) if d else ""


column_mapper_fn = Type[Callable[[Dict[str, str], Dict[str, str]], bool]]


def _get_validation_type(raw: dict):
    for k in "Validation Type", "Model":
        if k in raw:
            return raw[k]
    return ""


def map_dell_hostname(raw: dict, clean: dict) -> bool:
    hostname = raw.get("Host", "")
    name = hostname.lower().replace("_", "-")
    if "lf-" in name and not "-lf-" in name:
        name = name.replace("lf-", "-lf-")
    if "mxx-" in name:
        name = re.sub("mxx-", "mxxl-", name)
    if "-pd" in name:
        name = name.replace("pd", "pdu").replace("pdu-", "pdu")
    clean["name"] = name
    return bool(name)


def map_computacenter_location(raw: dict, clean: dict) -> bool:
    rawrack = raw.get("location.rack", raw.get("location rack ", raw.get("location rackÊ", "")))
    if "OOB2" in rawrack:
        rack = "inf006"  # random error
    else:
        rack = rawrack.split(" ")[-1].lower().replace("-", "")
    clean["location.rack"] = rack

    rawunit = raw.get("location.unit", "")
    try:
        unit = int(rawunit)
    except ValueError:
        unit = 0
    clean["location.unit"] = str(unit)

    clean["location.position"] = raw.get("location.position", "").lower()

    return bool(rack) and unit != ""


def map_computacenter_hostname(raw: dict, clean: dict) -> bool:
    rawname = raw.get("name", "").replace(" ", "-").replace(".", "-").strip().strip("-")

    if "7808" in raw.get("vendor.model", "").lower():
        rawname = "sp-sw01"

    name = rawname.split("/")[-1].split(" ")[-1].lower()

    name = name.replace(
        "-cmg-", "-mg-"
    ).replace("ztp-", "in-"
              ).replace("pd-01-", "pdu"
                        ).replace("pd-02-", "pdu"
                                  ).replace("kvu-", "kv-"
                                            ).replace("usr-", "us-"
                                                      ).replace("mc-", "mc")
    if name.startswith("cn-"):
        name = name.replace("cn-", "cn")

    # add rack if missing
    rack = clean["location.rack"]
    if not name.startswith(rack + "-"):
        name = f"{rack}-{name}"

    # add leading zero if missing
    m = re.match(".*[a-z](\d+)$", name)
    if m:
        suffix = m.group(1)
        name = name[:-len(suffix)] + f"{int(suffix):02}"

    clean["name"] = name
    return bool(name)


def map_computacenter_vendor_info(raw: dict, clean: dict) -> bool:
    rawvendor = raw.get("vendor.name", "").lower().replace(" ", "")
    rawmodel = raw.get("vendor.model", "").lower().strip()
    rawserial = raw.get("vendor.serial").lower().strip()

    if "r6615" in rawmodel:
        model = "r6615"
    else:
        model = rawmodel

    vendor = {
        "dell": "DL",
        "arista": "AR",
        "lenovo": "LN",
        "credo": "CR",
        "vertiv": "VE",
        "supermicro": "SM",
        "opengear": "OG",
        "": "",
    }.get(rawvendor, rawvendor)

    clean["vendor.name"] = vendor
    clean["vendor.serial"] = rawserial
    clean["vendor.model"] = model
    return bool(vendor)


def map_computacenter_mgmt_infos(raw: dict, clean: dict):
    dtype, drole, vendor = clean["type"], clean["role"], clean["vendor.name"]
    user, password, mac = "", "", format_mac(raw.get("management_info.mac", ""))
    if dtype == "SR":
        user, password = "root", "5cerebras"
    elif dtype == "SW":
        user, password = "admin", "CerebrasColo"
    elif dtype == "CN":
        user, password = "root", "CerebrasColo"
    elif dtype == "PR" and drole == "PU":
        if vendor == "EN":
            user, password = "admn", "admn"
    elif dtype == "PR":
        if vendor == "CR":
            user, password = "root", "0penBmc"
    clean["management_info.mac"] = mac
    clean["management_credentials.user"] = user
    clean["management_credentials.password"] = password
    # servers - we can discover mac through IPMI so not critical, but other devices, need MAC
    return True


def map_computacenter_ipmi_infos(raw: dict, clean: dict) -> bool:
    if clean["type"] != "SR":
        return True
    vendor, serial = clean["vendor.name"], clean["vendor.serial"]
    ipmi_name, ipmi_user, ipmi_password = "", raw.get("ipmi_credentials.user", ""), raw.get("ipmi_credentials.password",
                                                                                            "")
    if vendor == "DL":
        ipmi_name, ipmi_user = f"idrac-{serial}", "root"
    elif vendor == "HP":
        ipmi_name, ipmi_user = f"ILO{serial}", "Administrator"
    elif vendor == "SM":
        ipmi_name, ipmi_user, ipmi_password = clean[
                                                  "name"] + "-ipmi", "ADMIN", "Admin123" if not ipmi_password else ipmi_password
    clean["ipmi_credentials.user"] = ipmi_user
    clean["ipmi_credentials.password"] = ipmi_password
    clean["ipmi_info.name"] = ipmi_name
    clean["ipmi_info.mac"] = format_mac(raw.get("ipmi_info.mac", ""))
    return True


def map_computacenter_specific_fixes(raw: dict, clean: dict) -> bool:
    """ special hack mapper to fix issues with dell inventory specifications """

    # switch the ipmi / management mac for servers
    if clean["type"] == "SR" and "management_info.mac" in clean and "ipmi_info.mac" in clean:
        clean["management_info.mac"], clean["ipmi_info.mac"] = clean.get("ipmi_info.mac", ""), clean.get(
            "management_info.mac", "")

    if clean["type"] == "SW" and clean["role"] == "MG" and "wse" in clean["name"]:
        neighbor_rack_num = int(clean["location.rack"][3:]) - 1
        clean["switch_info.connected_racks"] = f"{clean['location.rack']},wse{neighbor_rack_num:03}"
        clean["subnet_info.prefixlen"] = '26'

    return True


def map_role(raw: dict, clean: dict) -> bool:
    hostname = clean["name"]
    role = ""
    if "-ax-" in hostname or "-ax_" in hostname or "-kv-" in hostname:
        role = "AX"
    if "-kvi-" in hostname:
        role = "IX"
    elif "-sx-" in hostname:
        role = "SX"
    elif "-ma-" in hostname:
        role = "MA"
    elif "-fe-" in hostname:
        role = "FE"
    elif "-mg-" in hostname or "-es-" in hostname:
        role = "MG"
    elif "lf-" in hostname:
        role = "LF"
    elif "-sp-" in hostname:
        role = "SP"
    elif "-mx-" in hostname or "-ml-" in hostname or "-mxx" in hostname:
        role = "MX"
    elif "-wk-" in hostname:
        role = "WK"
    elif "-us-" in hostname:
        role = "US"
    elif "-in-" in hostname or "-dcv-" in hostname:
        role = "IN"
    elif "-pa-pp" in hostname or "-mc" in hostname:
        role = "MC"  # credo media converter
    elif "-cn" in hostname:
        role = "IN"
    elif "-pd" in hostname:
        role = "PU"
    clean["role"] = role
    return bool(role)


def map_type(raw: dict, clean: dict) -> bool:
    hostname = clean["name"]
    t = ""
    if "-sr" in hostname or "_sr" in hostname:
        t = "SR"
    elif "-sw" in hostname:
        t = "SW"
    elif "-cn" in hostname:
        t = "CN"
    elif "-pa-pp" in hostname or "-mc" in hostname or "-pd" in hostname:
        t = "PR"
    clean["type"] = t
    return bool(t)


def map_dell_vendor_info(raw: dict, clean: dict) -> bool:
    role, t = clean["role"], clean["type"]
    vendor, serial = "", raw.get("Serial Number", "")
    validation_type = _get_validation_type(raw)
    model = raw.get("Model", "")
    if t == "SW":
        if role in ("LF", "SP"):
            vendor = "AR"
        else:
            if validation_type == "720DT-48S-2F":
                vendor = "AR"
            else:
                vendor = "DL"
    elif t == "SR":
        # validation type is the most accurate way to get vendor name
        if "R6615" in validation_type:
            vendor = "DL"
        elif "SM" in validation_type:
            vendor = "SM"
        else:
            # otherwise guess based on the role and convention
            if role in ["SX", "AX", "MX"]:
                if "mxxl" in clean["name"]:
                    vendor = "SM"
                else:
                    vendor = "DL"
            if role in ["MG", "WK"]:
                vendor = "HP"
            if role == "US":
                vendor = "SM"
    elif t == "CN":
        vendor = "OG"  # opengear
    elif t == "PU":
        if "6950" in (validation_type, model) or "6950PDU" in (validation_type, model):
            vendor = "EN"
    elif t == "PR":
        if role == "MC":
            vendor = "CR"
    clean["vendor.name"] = vendor
    clean["vendor.serial"] = serial
    return bool(vendor)


def map_dell_mgmt_infos(raw: dict, clean: dict):
    dtype, vendor = clean["type"], clean["vendor.name"]
    user, password, mac = "", "", raw.get("Management MAC", "").lower()
    if dtype == "SR":
        user, password = "root", "5cerebras"
        mac = raw.get("LOM 1 MAC", "").lower()
    elif dtype == "SW":
        user, password = "admin", "CerebrasColo"
    elif dtype == "CN":
        user, password = "root", "0penBmc"
    elif dtype == "PU":
        if vendor == "EN":
            user, password = "admn", "admn"
    elif dtype == "PR":
        if vendor == "CR":
            user, password = "admin", "CerebrasColo"
    clean["management_info.mac"] = mac
    clean["management_credentials.user"] = user
    clean["management_credentials.password"] = password
    # servers - we can discover mac through IPMI so not critical, but other devices, need MAC
    return True if dtype == "SR" else bool(mac)


def map_dell_ipmi_infos(raw: dict, clean: dict) -> bool:
    if clean["type"] != "SR":
        return True
    vendor, serial = clean["vendor.name"], clean["vendor.serial"]
    ipmi_name, ipmi_user, ipmi_password = "", "", raw.get("ILO Password", raw.get("ILO", ""))
    if vendor == "DL":
        ipmi_name, ipmi_user = f"idrac-{serial}", "root"
    elif vendor == "HP":
        ipmi_name, ipmi_user = f"ILO{serial}", "Administrator"
    elif vendor == "SM":
        ipmi_name, ipmi_user = clean["name"] + "-ipmi", "ADMIN"
    clean["ipmi_credentials.user"] = ipmi_user
    clean["ipmi_credentials.password"] = ipmi_password
    clean["ipmi_info.name"] = ipmi_name
    clean["ipmi_info.mac"] = raw.get("Management MAC", "").lower()
    return bool(ipmi_user) and bool(ipmi_password) and bool(clean["ipmi_info.mac"])


def map_dell_location(raw: dict, clean: dict) -> bool:
    rack = clean["name"].split("-")[0]
    clean["location.rack"] = rack
    clean["location.unit"] = raw.get("U Location", "0")
    return bool(rack)


def map_dell_specific_fixes(raw: dict, clean: dict) -> bool:
    """ special hack mapper to fix issues with dell inventory specifications """

    # bug in dell: reported LF switch as MG switches in certain cases
    validation_type = _get_validation_type(raw)
    if clean["name"].endswith("-mg-sw02") and "48S" not in validation_type:
        clean["name"] = clean["name"][:-len("-mg-sw02")] + "-lf-sw02"
        clean["role"] = "LF"
    if clean["name"].endswith("-mg-sw03") and "48S" not in validation_type:
        clean["name"] = clean["name"][:-len("-mg-sw03")] + "-lf-sw03"
        clean["role"] = "LF"

    return True


def map_user_input_properties(raw: dict, clean: dict) -> bool:
    """ special hack mapper to put in user defined properties - raw here is not the excel file input but user input """
    for k, v in raw.items():
        if k == "switch_info.deploy_subnet":
            if "switch_info.deploy_subnet" in raw and clean["type"] == "SW" and clean["role"] == "MG":
                clean["switch_info.deploy_subnet"] = raw["switch_info.deploy_subnet"]
        else:
            clean[k] = v

    return True


def computacenter_ignorer(raw: dict) -> bool:
    if "VR3357SPBDL" in raw.get("vendor.model", "") or "VR3307SPBDL" in raw.get("vendor.model", ""):  # rack records
        return True
    if "DCS-7800R4C-36PE" in raw.get("vendor.model", ""):  # cable bundles
        return True
    if "RESERVE" == raw.get("vendor.model", ""):
        return True
    return False


def parse_file(file_path: str, global_props: dict = None, parser_type="dell") -> List[dict]:
    if parser_type == "dell":
        ignorer = lambda raw: False
        mappers: Dict[str, column_mapper_fn] = {fn.__name__: fn for fn in [
            map_dell_hostname,
            map_type,
            map_role,
            map_dell_vendor_info,
            map_dell_mgmt_infos,
            map_dell_ipmi_infos,
            map_dell_location,
            map_dell_specific_fixes,
        ]}
    elif parser_type == "computacenter":
        ignorer = computacenter_ignorer
        mappers: Dict[str, column_mapper_fn] = {fn.__name__: fn for fn in [
            map_computacenter_location,
            map_computacenter_hostname,
            map_type,
            map_role,
            map_computacenter_vendor_info,
            map_computacenter_mgmt_infos,
            map_computacenter_ipmi_infos,
            map_computacenter_specific_fixes,
        ]}
    else:
        raise AssertionError(f"unknown parser type={parser_type}")

    rows = []
    with open(file_path, 'r') as f:
        reader = csv.DictReader(f)
        for d in reader:
            if not d or ignorer(d):
                continue
            cleaned_device = {}
            for mapper_name, mapper in mappers.items():
                if not mapper(d, cleaned_device):
                    error = [f"{file_path} line {reader.line_num}", f"failed mapper '{mapper_name}'",
                             f"input:\n{fmt_dict(d)}", f"output so far:\n{fmt_dict(cleaned_device)}"]
                    cleaned_device["_error"] = error
                    rows.append(cleaned_device)
                    break
            else:
                if global_props:
                    map_user_input_properties(global_props, cleaned_device)
                rows.append(cleaned_device)
    return rows


validation_always_required = ["location.rack", "vendor.name"]
validation_required_fields = {
    "SR": [
        *validation_always_required,
        "management_credentials.user",
        "management_credentials.password",
        "ipmi_info.mac",
        "ipmi_credentials.user",
        "ipmi_credentials.password",
    ],
    "SW": [
        *validation_always_required,
        "management_info.mac",
        "management_credentials.user",
        "management_credentials.password",
    ],
    "PR": [
        *validation_always_required,
        "management_info.mac",
        "management_credentials.user",
        "management_credentials.password",
    ],
}


def validate_entry(e: dict) -> List[str]:
    """ Simple validations - the import_inventory call also does validation so this is a first alert """
    type = e['type']
    req_fields = validation_required_fields[type] if type in validation_required_fields else validation_always_required
    missing = []
    for field in req_fields:
        if not e.get(field):
            missing.append(field)
    if missing:
        return [f"missing required field(s) from {e['name']}: {','.join(missing)}"]
    return []


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Parses the given rack integrator xlsx or csv files and outputs a CSV of the normalized inventory of all the files."
    )
    parser.add_argument(
        "files",
        metavar="FILE",
        type=str,
        nargs="+",
        help="One or more asset files to parse and normalize."
    )
    parser.add_argument(
        "--format",
        choices=("dell", "computacenter"),
        default="dell",
        help="Rack integrator format. Each integrator provides different semantics so parse accordingly"
    )
    parser.add_argument(
        "--output",
        metavar="OUTPUT",
        type=str,
        default="-",
        help="Path to the output file. Use '-' to output to stdout (default: '-')."
    )
    parser.add_argument(
        "--deploy_subnet",
        metavar="DEPLOY_SUBNET",
        type=str,
        default="",
        help="Optional deployment subnet to assign to the mg switches."
    )
    parser.add_argument(
        "--properties",
        metavar="PROPERTIES",
        type=str,
        nargs="+",
        default="",
        help="Properties to set on all devices. Useful to set deploy.tag or location.stamp"
    )
    parser.add_argument(
        "--errors",
        choices=("include", "exclude", "only"),
        type=str,
        default="exclude",
        help="Whether to include devices with errors, exclude them, or only display them"
    )

    args = parser.parse_args()
    global_props = {}
    for v in args.properties:
        k, v = v.split("=", 1)
        global_props[k] = v
    if args.deploy_subnet:
        global_props["switch_info.deploy_subnet"] = args.deploy_subnet

    output = []
    for file in args.files:
        try:
            if file.endswith(".xlsx"):
                csv_file = convert_xlsx_to_csv(file, start_marker='U Location' if args.format == "dell" else None)
            elif not file.endswith(".csv"):
                logger.warning(f"ignoring file {file}: not .csv or .xlsx")
                continue
            else:
                csv_file = file
            output.extend(parse_file(csv_file, global_props, parser_type=args.format))
        except Exception as e:
            logger.exception(e)
            logger.error(f"failed to parse {file}: {e}")

    output = sorted(output, key=lambda e: e.get("name", ""))
    for e in output:
        if "_error" not in e:
            e["_error"] = validate_entry(e)

    # add empty string as default value for all devices
    preferred_column_order = [
        "name",
        "_error",
        "role",
        "type",
        "location.rack",
        "location.unit",
        "vendor.name",
        "vendor.serial",
        "management_info.mac",
        "management_credentials.user",
        "management_credentials.password",
        "ipmi_info.mac",
        "ipmi_info.name",
        "ipmi_credentials.user",
        "ipmi_credentials.password",
    ]

    column_names = set()
    for device in output:
        column_names.update(device.keys())
    col_sort_order = sorted([n for n in list(column_names) if n not in preferred_column_order])
    column_names = sorted(list(column_names),
                          key=lambda n: preferred_column_order.index(n) if n in preferred_column_order else len(
                              preferred_column_order) + col_sort_order.index(n))
    for device in output:
        for cname in column_names:
            if not cname in device:
                device[cname] = ""

    # print a summary of what was parsed
    rack_sums = defaultdict(int)
    error_devices = []
    for e in output:
        if e.get("_error"):
            error_devices.append(e)
        else:
            rack_sums[e["location.rack"] + ":" + e["type"]] += 1
            rack_sums[e["location.rack"] + ":" + e["type"] + ":" + e["role"]] += 1
    for k in sorted(rack_sums.keys()):
        logger.info(f"{k}: {rack_sums[k]}")
    if error_devices:
        for e in error_devices:
            logger.error("\n".join(e['_error']))
        logger.error(f"{len(error_devices)} total errors")

    print(column_names)

    of, close = sys.stdout, lambda: ""
    if args.output != "-":
        of = open(args.output, 'w')
        close = of.close

    if args.errors == "exclude" and "_error" in column_names:
        column_names.pop(column_names.index("_error"))

    writer = csv.DictWriter(of, fieldnames=column_names)
    writer.writeheader()
    for e in output:

        if "_error" in e:
            e["_error"] = (e["_error"] if isinstance(e["_error"], str) else "; ".join(e["_error"])).replace("\n", ";")

        if args.errors == "include":
            writer.writerow(e)
        elif args.errors == "exclude" and ("_error" not in e or not bool(e.get("_error"))):
            del e["_error"]
            writer.writerow(e)
        elif args.errors == "only" and bool(e.get("_error")):
            writer.writerow(e)

    close()

    if error_devices:
        exit(1)
